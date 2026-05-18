import os
from flask import Flask, render_template, request, redirect, url_for, send_file, Response
import csv
import io
import threading
import uuid
from models import db, Scan, Event
from scraper import scrape_forex_calendar

app = Flask(__name__)

# Database Configuration
app.config['SECRET_KEY'] = 'your-super-secret-key-change-in-production'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///intelligence.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db.init_app(app)

with app.app_context():
    db.create_all()

# In-memory store to track the background thread progress (fully zero-install & highly reliable)
PROGRESS_STORE = {}

def run_background_scrape(task_id, start_date, end_date, currencies, impacts, asset_type, scan_id):
    PROGRESS_STORE[task_id] = {
        'state': 'PROGRESS',
        'percent': 0,
        'status': 'Initializing synchronization...',
        'scan_id': scan_id
    }
    
    def update_progress(percent, status):
        PROGRESS_STORE[task_id] = {
            'state': 'PROGRESS',
            'percent': percent,
            'status': status,
            'scan_id': scan_id
        }

    try:
        # Flask requires application context to perform DB queries in a background thread
        with app.app_context():
            scrape_forex_calendar(
                start_date=start_date, 
                end_date=end_date, 
                currencies=currencies, 
                impacts=impacts, 
                asset_type=asset_type,
                scan_id=scan_id,
                db_session=db.session,
                progress_callback=update_progress
            )
        
        PROGRESS_STORE[task_id] = {
            'state': 'SUCCESS',
            'percent': 100,
            'status': 'Completed',
            'scan_id': scan_id
        }
    except Exception as e:
        PROGRESS_STORE[task_id] = {
            'state': 'FAILURE',
            'percent': 0,
            'status': 'Synchronization Failed',
            'error': str(e),
            'scan_id': scan_id
        }

@app.route('/')
def index():
    return render_template('Index.html')

@app.route('/run', methods=['POST'])
def run():
    scan_name = request.form.get("scan_name", "Latest_Sync")
    start_date = request.form.get("start_date")
    end_date = request.form.get("end_date")
    asset_type = request.form.get("asset_type", "currency")
    impacts = request.form.getlist("impact")
    currencies = request.form.getlist("currencies") # Dynamic items from UI
    
    new_scan = Scan(name=scan_name, start_date=start_date, end_date=end_date, asset_type=asset_type)
    db.session.add(new_scan)
    db.session.commit()
    
    # Generate a unique task ID for this background operation
    task_id = str(uuid.uuid4())
    
    # Start standard Python background thread for complete Zero-Install reliability on Windows
    thread = threading.Thread(
        target=run_background_scrape,
        args=(task_id, start_date, end_date, currencies, impacts, asset_type, new_scan.id)
    )
    thread.daemon = True # Close thread when Flask app closes
    thread.start()
    
    # Send user to the loading screen tracking this thread
    return redirect(url_for('loading', task_id=task_id, scan_id=new_scan.id))

@app.route('/loading/<task_id>/<int:scan_id>')
def loading(task_id, scan_id):
    return render_template('loading.html', task_id=task_id, scan_id=scan_id)

@app.route('/status/<task_id>')
def task_status(task_id):
    task = PROGRESS_STORE.get(task_id)
    if not task:
        return {'state': 'PENDING', 'status': 'Initializing thread...', 'percent': 0}
        
    response = {'state': task['state']}
    
    if task['state'] == 'PROGRESS':
        response.update({
            'status': task['status'],
            'percent': task['percent'],
            'scan_id': task['scan_id']
        })
    elif task['state'] == 'SUCCESS':
        response.update({
            'status': 'Completed',
            'percent': 100,
            'scan_id': task['scan_id']
        })
    elif task['state'] == 'FAILURE':
        response.update({
            'status': task['status'],
            'percent': 0,
            'error': task.get('error', 'Unknown Error'),
            'scan_id': task['scan_id']
        })
    else:
        response.update({
            'status': 'Pending...',
            'percent': 0
        })
        
    return response

@app.route('/dashboard')
def dashboard():
    all_scans = Scan.query.order_by(Scan.timestamp.desc()).all()
    scan_id = request.args.get('scan_id', type=int)
    selected_scan = None
    
    if scan_id:
        selected_scan = db.session.get(Scan, scan_id)
    elif all_scans:
        selected_scan = all_scans[0]
    
    leads = selected_scan.events if selected_scan else []
    
    volatility_data = {}
    if leads:
        for event in leads:
            curr = event.currency
            if curr not in volatility_data:
                volatility_data[curr] = {'High': 0, 'Medium': 0, 'Low': 0, 'score': 0}
            
            impact = event.impact
            volatility_data[curr][impact] += 1
            
            if impact == 'High': volatility_data[curr]['score'] += 3
            elif impact == 'Medium': volatility_data[curr]['score'] += 1
    
    sorted_volatility = dict(sorted(volatility_data.items(), key=lambda item: item[1]['score'], reverse=True))

    return render_template('dashboard.html', 
                           leads=leads, 
                           all_scans=all_scans, 
                           volatility=sorted_volatility,
                           current_scan_id=selected_scan.id if selected_scan else None,
                           current_scan_name=selected_scan.name if selected_scan else "No Active Sync")

@app.route('/download/<int:scan_id>')
def download(scan_id):
    import pandas as pd
    from openpyxl.styles import Font, Alignment, PatternFill
    
    scan = db.get_or_404(Scan, scan_id)
    
    # Structure events into a list of dictionaries for pandas DataFrame
    events_data = []
    for event in scan.events:
        events_data.append({
            'Date': event.date,
            'Time': event.time,
            'Currency': event.currency,
            'Impact': event.impact,
            'Event Description': event.event_title,
            'Killzone': event.killzone,
            'NLP Tags': event.tags
        })
    
    df = pd.DataFrame(events_data)
    
    # Handle edge case where the scan contains no events
    if df.empty:
        df = pd.DataFrame(columns=['Date', 'Time', 'Currency', 'Impact', 'Event Description', 'Killzone', 'NLP Tags'])
        
    # Write to an in-memory byte buffer
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Market Intelligence')
        
        workbook = writer.book
        worksheet = writer.sheets['Market Intelligence']
        
        # Style headers to match the DataPulse Premium look (Espresso Brown Fill with white text)
        header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='2A1E17', end_color='2A1E17', fill_type='solid') # Luxury Espresso
        header_align = Alignment(horizontal='center', vertical='center')
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            
        # Style data cells (Outfit/Segoe UI, clean alignments, and soft grid structures)
        data_font = Font(name='Segoe UI', size=10)
        data_align = Alignment(horizontal='left', vertical='center')
        center_align = Alignment(horizontal='center', vertical='center')
        
        # Apply alignment across data row matrices
        for row in range(2, worksheet.max_row + 1):
            for col in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.font = data_font
                # Center alignments for short-code metadata columns
                if col in [1, 2, 3, 4, 6]:
                    cell.alignment = center_align
                else:
                    cell.alignment = data_align
                    
        # Set premium double height for the header row
        worksheet.row_dimensions[1].height = 28
        
        # Dynamically auto-fit column widths based on text length to prevent truncating text
        for col in worksheet.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                val = str(cell.value or '')
                # Filter long paragraphs to prevent excessive column stretching
                if len(val) < 60:
                    max_len = max(max_len, len(val))
            # Auto-set width with default minimal safety padding
            worksheet.column_dimensions[col_letter].width = max(max_len + 4, 12)
            
    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-disposition": f"attachment; filename={scan.name.replace(' ', '_')}.xlsx"}
    )

@app.route('/seed')
def seed():
    # Clear existing scans and events first
    Scan.query.delete()
    Event.query.delete()
    db.session.commit()
    
    # Create a beautiful scan
    demo_scan = Scan(
        name="Global Multi-Asset Sync (Demo)", 
        start_date="2026-05-18", 
        end_date="2026-05-25", 
        asset_type="currency"
    )
    db.session.add(demo_scan)
    db.session.commit()
    
    # Add beautiful events
    events = [
        Event(
            scan_id=demo_scan.id,
            date="May 18, 2026",
            time="08:30 AM",
            currency="USD",
            impact="High",
            event_title="Core CPI m/m",
            explanation="<strong>Market Impact Analysis:</strong> Consumer Price Index measures the change in the price of goods and services from the perspective of the consumer. This is a critical indicator of inflation, which heavily influences Federal Reserve monetary policy. A higher-than-expected reading is bullish for the USD, as it signals potential rate hikes, whereas a lower reading suggests inflationary cooling.",
            killzone="NY Open",
            tags="CPI,Inflation,Fed Policy"
        ),
        Event(
            scan_id=demo_scan.id,
            date="May 18, 2026",
            time="11:45 AM",
            currency="EUR",
            impact="High",
            event_title="ECB President Lagarde Speaks",
            explanation="<strong>Market Impact Analysis:</strong> President Christine Lagarde's public statements often dictate short-term Euro volatility. Analysts will inspect her tone for hawkish or dovish signals regarding the upcoming Governing Council meeting. Increased focus is expected on the ECB's balance sheet reduction progress.",
            killzone="London Close",
            tags="ECB,Monetary Policy,Lagarde"
        ),
        Event(
            scan_id=demo_scan.id,
            date="May 19, 2026",
            time="02:00 PM",
            currency="USD",
            impact="High",
            event_title="FOMC Interest Rate Decision",
            explanation="<strong>Market Impact Analysis:</strong> The Federal Open Market Committee determines global interest rate benchmarks. This event is the single most volatile catalyst in global financial markets. Alongside the rate announcement, the FOMC statement and economic projections (Dot Plot) will indicate policy trajectory for the rest of 2026.",
            killzone="NY Midday",
            tags="FOMC,Fed Rates,Quantitative Tightening"
        ),
        Event(
            scan_id=demo_scan.id,
            date="May 19, 2026",
            time="02:30 PM",
            currency="USD",
            impact="High",
            event_title="FOMC Press Conference",
            explanation="<strong>Market Impact Analysis:</strong> Fed Chairman Powell's address directly following the rate decision provides key insights into the committee's consensus. Extreme volatility is anticipated during the live Q&A session as traders interpret forward guidance statements.",
            killzone="NY Midday",
            tags="Fed,Powell,Forward Guidance"
        ),
        Event(
            scan_id=demo_scan.id,
            date="May 20, 2026",
            time="04:30 AM",
            currency="GBP",
            impact="High",
            event_title="CPI y/y",
            explanation="<strong>Market Impact Analysis:</strong> Year-over-year CPI data for the United Kingdom. With the Bank of England maintaining a cautious stance, CPI deviation from the target will dictate whether the MPC accelerates or halts the current monetary easing cycle.",
            killzone="London Open",
            tags="GBP,UK Inflation,BoE"
        ),
        Event(
            scan_id=demo_scan.id,
            date="May 20, 2026",
            time="08:30 AM",
            currency="CAD",
            impact="Medium",
            event_title="Retail Sales m/m",
            explanation="<strong>Market Impact Analysis:</strong> Retail Sales represent the primary gauge of consumer spending, driving the majority of economic activity in Canada. Stronger sales reflect consumer resilience, bolstering the Bank of Canada's case for restrictive rates.",
            killzone="NY Open",
            tags="CAD,Retail,Consumer Strength"
        ),
        Event(
            scan_id=demo_scan.id,
            date="May 21, 2026",
            time="09:15 PM",
            currency="AUD",
            impact="Medium",
            event_title="RBA Gov Bullock Speaks",
            explanation="<strong>Market Impact Analysis:</strong> Governor Bullock's speech at the Sydney Banking Summit will likely touch on domestic employment growth and commodity export price dynamics, impacting the AUD/USD pair.",
            killzone="Tokyo Open",
            tags="AUD,RBA,Bullock"
        ),
        Event(
            scan_id=demo_scan.id,
            date="May 22, 2026",
            time="10:00 AM",
            currency="XAU",
            impact="High",
            event_title="Geopolitical Risk Premium Surge",
            explanation="<strong>Market Impact Analysis:</strong> Macro indicators show capital fleeing standard fiat corridors into gold contracts as systemic credit risks rise. Standard USD correlations have decoupled, highlighting safe-haven flows.",
            killzone="London/NY Overlap",
            tags="Safe Haven,Gold,Geopolitics"
        ),
        Event(
            scan_id=demo_scan.id,
            date="May 22, 2026",
            time="11:30 AM",
            currency="BTC",
            impact="High",
            event_title="Global Dollar Liquidity Shift",
            explanation="<strong>Market Impact Analysis:</strong> Federal Reserve net liquidity injections trigger a structural expansion in digital asset protocols. BTC open interest on CME reaches record levels as institutional volume increases.",
            killzone="NY Open",
            tags="BTC,Liquidity,Crypto"
        )
    ]
    
    for event in events:
        db.session.add(event)
        
    db.session.commit()
    return redirect(url_for('dashboard', scan_id=demo_scan.id))

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host='0.0.0.0', port=port, debug=False)